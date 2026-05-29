import asyncio
import io
import json
import os
import re
import subprocess
import time
from typing import Any
import httpx
from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from langchain_core.chat_history import InMemoryChatMessageHistory
from langchain_core.messages import AIMessage, HumanMessage
from openpyxl import Workbook
from pydantic import BaseModel, Field
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
load_dotenv(os.path.join(BASE_DIR, ".env"))
load_dotenv(os.path.join(os.path.dirname(BASE_DIR), ".env"), override=False)

VALID_MODES = {"investor", "business_model", "customer", "campaign_builder", "financial"}
TOKEN_LIMIT = int(os.getenv("PRO_TOKEN_LIMIT", "150000"))
TOKEN_WINDOW_SECONDS = 24 * 60 * 60
SESSION_TTL_SECONDS = 2 * 60 * 60  # 2-hour session expiry
SIDECAR_URL = os.getenv("NODE_SIDECAR_URL", "http://127.0.0.1:3000")

app = FastAPI(
    title="Jeff AI Agent API",
    description="FastAPI endpoint wrapping the persistent TypeScript Jeff agent sidecar.",
    version="1.0.0",
)

cors_origins = [
    origin.strip()
    for origin in os.getenv(
        "CORS_ORIGINS",
        "https://juststrtup.com,http://juststrtup.com,https://www.juststrtup.com,https://juststartup.com,http://juststartup.com,https://www.juststartup.com,http://localhost:8080,http://localhost:8000,http://127.0.0.1:8080,http://127.0.0.1:8000",
    ).split(",")
    if origin.strip()
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=cors_origins,
    allow_credentials=True,
    allow_methods=["GET", "POST", "OPTIONS"],
    allow_headers=["*"],
    expose_headers=[
        "X-Session-Id",
        "X-Tokens-Limit",
        "X-Tokens-Remaining",
        "X-Tokens-Reset",
        "Retry-After",
    ],
)

node_process: subprocess.Popen | None = None
# TODO: replace dict with Redis for multi-instance production
session_store: dict[str, InMemoryChatMessageHistory] = {}
session_timestamps: dict[str, float] = {}  # last-access time per session
# TODO: replace dict with Redis for multi-instance production
token_usage_store: dict[str, list[tuple[float, int]]] = {}


class ChatRequest(BaseModel):
    message: str = Field(..., min_length=1)
    mode: str
    session_id: str = "default_session"
    user_id: str | None = None


class ExportRequest(BaseModel):
    payload: Any | None = None
    content: Any | None = None
    filename: str = "jeff-export"
    title: str | None = None


async def _sidecar_is_ready() -> bool:
    try:
        async with httpx.AsyncClient(timeout=1.0) as client:
            response = await client.get(f"{SIDECAR_URL}/health")
        return response.status_code == 200
    except Exception:
        return False


async def _wait_for_sidecar(timeout_seconds: int = 30) -> None:
    deadline = time.time() + timeout_seconds
    while time.time() < deadline:
        if await _sidecar_is_ready():
            return
        if node_process and node_process.poll() is not None:
            raise RuntimeError(f"Node sidecar exited early with code {node_process.returncode}")
        await asyncio.sleep(0.5)
    raise RuntimeError(f"Node sidecar did not become healthy at {SIDECAR_URL}")


@app.on_event("startup")
async def startup_event() -> None:
    global node_process

    if os.getenv("START_NODE_SIDECAR", "true").lower() == "false":
        return
    if await _sidecar_is_ready():
        return

    node_cmd = "node.exe" if os.name == "nt" else "node"
    compiled_server = os.path.join(BASE_DIR, "dist", "server.js")

    if not os.path.exists(compiled_server):
        raise RuntimeError(
            f"Compiled sidecar not found at {compiled_server}. "
            "Run 'npm run build' before starting the server."
        )

    node_process = subprocess.Popen([node_cmd, compiled_server], cwd=BASE_DIR)
    await _wait_for_sidecar()


@app.on_event("shutdown")
def shutdown_event() -> None:
    if node_process and node_process.poll() is None:
        node_process.terminate()


def get_session_history(session_id: str) -> InMemoryChatMessageHistory:
    now = time.time()
    # Expire sessions older than SESSION_TTL_SECONDS (2 hours)
    last_access = session_timestamps.get(session_id, 0)
    if session_id in session_store and (now - last_access) > SESSION_TTL_SECONDS:
        del session_store[session_id]
        del session_timestamps[session_id]
    if session_id not in session_store:
        session_store[session_id] = InMemoryChatMessageHistory()
    session_timestamps[session_id] = now
    return session_store[session_id]


def quota_key_for(request: ChatRequest) -> str:
    return request.user_id or request.session_id


def prune_token_events(key: str, now: float | None = None) -> list[tuple[float, int]]:
    now = now or time.time()
    cutoff = now - TOKEN_WINDOW_SECONDS
    events = [(created_at, tokens) for created_at, tokens in token_usage_store.get(key, []) if created_at > cutoff]
    token_usage_store[key] = events
    return events


def quota_state(key: str) -> dict[str, int]:
    now = time.time()
    events = prune_token_events(key, now)
    used = sum(tokens for _, tokens in events)
    reset_at = int((events[0][0] + TOKEN_WINDOW_SECONDS) if events else (now + TOKEN_WINDOW_SECONDS))
    remaining = max(0, TOKEN_LIMIT - used)
    return {"used": used, "remaining": remaining, "reset_at": reset_at}


def record_token_usage(key: str, input_tokens: int, output_tokens: int) -> None:
    total = max(0, int(input_tokens or 0)) + max(0, int(output_tokens or 0))
    if total <= 0:
        return
    prune_token_events(key)
    token_usage_store.setdefault(key, []).append((time.time(), total))


def quota_headers(session_id: str, key: str) -> dict[str, str]:
    state = quota_state(key)
    return {
        "X-Session-Id": session_id,
        "X-Tokens-Limit": str(TOKEN_LIMIT),
        "X-Tokens-Remaining": str(state["remaining"]),
        "X-Tokens-Reset": str(state["reset_at"]),
    }


def agent_history_from_session(chat_history: InMemoryChatMessageHistory) -> list[dict[str, Any]]:
    messages: list[dict[str, Any]] = []
    for msg in chat_history.messages:
        if isinstance(msg, HumanMessage):
            messages.append({"role": "user", "content": [{"type": "input_text", "text": msg.content}]})
        elif isinstance(msg, AIMessage):
            messages.append({"role": "assistant", "content": [{"type": "output_text", "text": msg.content}]})
    return messages


def safe_filename(filename: str, extension: str) -> str:
    base = re.sub(r"[^A-Za-z0-9_.-]+", "-", filename).strip(".-") or "jeff-export"
    if not base.lower().endswith(f".{extension}"):
        base = f"{base}.{extension}"
    return base


def export_payload(request: ExportRequest) -> Any:
    payload = request.payload if request.payload is not None else request.content
    if payload is None:
        raise HTTPException(status_code=422, detail="Either payload or content is required.")
    return payload


def scalar(value: Any) -> bool:
    return value is None or isinstance(value, (str, int, float, bool))


def write_table(ws, rows: list[Any], start_row: int) -> int:
    if not rows:
        return start_row
    if all(isinstance(row, dict) for row in rows):
        headers = sorted({key for row in rows for key in row.keys()})
        for col, header in enumerate(headers, 1):
            ws.cell(start_row, col, header)
        for row_index, row in enumerate(rows, start_row + 1):
            for col, header in enumerate(headers, 1):
                value = row.get(header)
                ws.cell(row_index, col, value if scalar(value) else json.dumps(value, ensure_ascii=False))
        return start_row + len(rows) + 2
    if all(isinstance(row, list) for row in rows):
        for row_index, row in enumerate(rows, start_row):
            for col, value in enumerate(row, 1):
                ws.cell(row_index, col, value if scalar(value) else json.dumps(value, ensure_ascii=False))
        return start_row + len(rows) + 1
    for offset, value in enumerate(rows):
        ws.cell(start_row + offset, 1, value if scalar(value) else json.dumps(value, ensure_ascii=False))
    return start_row + len(rows) + 1


def populate_workbook(ws, payload: Any) -> None:
    if isinstance(payload, list):
        write_table(ws, payload, 1)
        return
    if isinstance(payload, dict):
        row = 1
        for key, value in payload.items():
            ws.cell(row, 1, key)
            if isinstance(value, list):
                row += 1
                row = write_table(ws, value, row)
            elif isinstance(value, dict):
                ws.cell(row, 2, json.dumps(value, ensure_ascii=False, indent=2))
                row += 1
            else:
                ws.cell(row, 2, value)
                row += 1
        return
    for row, line in enumerate(str(payload).splitlines() or [str(payload)], 1):
        ws.cell(row, 1, line)


def pdf_elements(payload: Any, title: str) -> list[Any]:
    styles = getSampleStyleSheet()
    elements: list[Any] = [Paragraph(title, styles["Title"]), Spacer(1, 12)]

    if isinstance(payload, list) and payload and all(isinstance(row, dict) for row in payload):
        headers = sorted({key for row in payload for key in row.keys()})
        data = [headers] + [[str(row.get(header, "")) for header in headers] for row in payload]
        table = Table(data, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eeeeee")),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        elements.append(table)
        return elements

    text = payload if isinstance(payload, str) else json.dumps(payload, ensure_ascii=False, indent=2)
    for line in str(text).splitlines() or [str(text)]:
        elements.append(Paragraph(line.replace(" ", "&nbsp;"), styles["BodyText"]))
    return elements


@app.post("/chat")
async def chat(request: ChatRequest):
    if request.mode not in VALID_MODES:
        raise HTTPException(
            status_code=422,
            detail=f"Invalid mode '{request.mode}'. Must be one of: {', '.join(sorted(VALID_MODES))}",
        )

    quota_key = quota_key_for(request)
    current_quota = quota_state(quota_key)
    if current_quota["remaining"] <= 0:
        retry_after = max(1, current_quota["reset_at"] - int(time.time()))
        raise HTTPException(
            status_code=429,
            detail="Token limit exceeded for the last 24 hours.",
            headers={
                "Retry-After": str(retry_after),
                "X-Tokens-Limit": str(TOKEN_LIMIT),
                "X-Tokens-Remaining": "0",
                "X-Tokens-Reset": str(current_quota["reset_at"]),
            },
        )

    chat_history = get_session_history(request.session_id)
    messages = agent_history_from_session(chat_history)
    messages.append({"role": "user", "content": [{"type": "input_text", "text": request.message}]})

    payload = {"messages": messages, "mode": request.mode}

    async def stream_response():
        full_response = ""
        input_tokens = 0
        output_tokens = 0
        try:
            async with httpx.AsyncClient(timeout=None) as client:
                async with client.stream("POST", f"{SIDECAR_URL}/chat", json=payload) as sidecar_response:
                    if sidecar_response.status_code != 200:
                        error_text = await sidecar_response.aread()
                        yield f"Jeff is unavailable right now. Sidecar returned {sidecar_response.status_code}: {error_text.decode(errors='replace')}"
                        return

                    async for line in sidecar_response.aiter_lines():
                        if not line:
                            continue
                        try:
                            event = json.loads(line)
                        except json.JSONDecodeError:
                            continue

                        event_type = event.get("type")
                        if event_type == "token":
                            text = event.get("text", "")
                            full_response += text
                            yield text
                        elif event_type == "usage":
                            usage = event.get("usage") or {}
                            input_tokens = int(usage.get("inputTokens") or usage.get("input_tokens") or 0)
                            output_tokens = int(usage.get("outputTokens") or usage.get("output_tokens") or 0)
                        elif event_type == "error":
                            yield event.get("message", "Jeff hit an internal error.")

            if full_response:
                chat_history.add_message(HumanMessage(content=request.message))
                chat_history.add_message(AIMessage(content=full_response))
            record_token_usage(quota_key, input_tokens, output_tokens)
        except Exception as exc:
            yield f"Stream failed: {exc}"

    headers = {
        "Cache-Control": "no-cache",
        "X-Accel-Buffering": "no",
        **quota_headers(request.session_id, quota_key),
    }
    return StreamingResponse(stream_response(), media_type="text/plain", headers=headers)


@app.post("/export/xlsx")
async def export_xlsx(request: ExportRequest):
    payload = export_payload(request)
    wb = Workbook()
    ws = wb.active
    ws.title = "Jeff Export"
    populate_workbook(ws, payload)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = safe_filename(request.filename, "xlsx")
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/export/pdf")
async def export_pdf(request: ExportRequest):
    payload = export_payload(request)
    output = io.BytesIO()
    title = request.title or request.filename or "Jeff Export"
    document = SimpleDocTemplate(output, pagesize=letter, title=title)
    document.build(pdf_elements(payload, title))
    output.seek(0)

    filename = safe_filename(request.filename, "pdf")
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/clear")
async def clear_session(request: Request):
    data = await request.json()
    session_id = data.get("session_id")
    if session_id in session_store:
        del session_store[session_id]
    return {"status": "cleared", "session_id": session_id}


@app.get("/health")
async def health_check():
    sidecar_ready = await _sidecar_is_ready()
    return {"status": "ok", "service": "Jeff AI Agent", "sidecar": "ok" if sidecar_ready else "unavailable"}


@app.get("/")
async def serve_frontend():
    ui_path = os.path.join(BASE_DIR, "jeff-ui.html")
    if os.path.exists(ui_path):
        return FileResponse(ui_path)
    return {"message": "Jeff AI Agent API is running."}
